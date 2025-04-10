import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def save_to_excel(data, file_path):
    """Save parsed data to an Excel file with formatting."""
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active

    # Apply styles to headers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, ws.max_row + 1):
            cell = ws[f"C{row}"]
            if cell.value:
                cell.value = f'=HYPERLINK("{cell.value}", "Ссылка")'
                cell.style = "Hyperlink"

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter != "C":
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)